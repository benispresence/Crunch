"""
Recursive folder scanner for the File connection type.

Walks a local directory (or single file) and classifies each entry by
its leaf extension. Skips anything we don't recognise so the user
sees a clean list. For Excel workbooks the scanner expands each sheet
into its own entry so they can be selected individually — Excel
workbooks frequently hold three or four logically separate tables.

The scanner is local-only: cloud bucket walks (s3 / gs / azure) need
provider SDKs and are deliberately left out of this first cut. The
``database`` field on a file connection already accepts a cloud URI;
this scanner is purely an aid for "I have a folder of mixed files,
which ones do I want?".
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path

from crunch.connections.adapters.file_adapter import (
    _extension_of,
    ARROW_EXTENSIONS,
    CSV_EXTENSIONS,
    EXCEL_EXTENSIONS,
    JSON_EXTENSIONS,
    PARQUET_EXTENSIONS,
)

logger = logging.getLogger(__name__)


def _format_label(ext: str) -> str:
    if ext in CSV_EXTENSIONS:
        return "csv"
    if ext in PARQUET_EXTENSIONS:
        return "parquet"
    if ext in JSON_EXTENSIONS:
        return "json"
    if ext in ARROW_EXTENSIONS:
        return "arrow"
    if ext in EXCEL_EXTENSIONS:
        return "excel"
    return "unknown"


@dataclass
class ScannedFile:
    """One row in the scan result. ``uri`` is what the FileAdapter
    consumes — either an absolute path or, for an Excel sheet, the
    path with a ``#SheetName`` suffix so the adapter knows which
    sheet to register."""

    uri: str
    name: str
    format: str
    size_bytes: int
    relative_path: str
    sheet: str | None = None


@dataclass
class ScanResult:
    root: str
    files: list[ScannedFile] = field(default_factory=list)
    skipped: int = 0
    error: str | None = None


def list_excel_sheets(path: Path) -> list[str]:
    """Return the sheet names in an Excel workbook, or an empty list
    if the workbook can't be opened (corrupt, password-protected, etc.).
    Uses openpyxl's read-only mode to avoid loading cell data."""
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        return []
    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    except Exception as e:
        logger.warning("Could not read sheets from %s: %s", path, e)
        return []


def scan_folder(
    root: str,
    *,
    recursive: bool = True,
    max_files: int = 5000,
) -> ScanResult:
    """Walk ``root`` and return a :class:`ScanResult` describing every
    supported file underneath. ``max_files`` caps the result so a
    pathological directory tree doesn't blow up the UI.

    Honours ``NICEMETA_FILE_SOURCE_ROOT`` (same env var as
    :mod:`file_adapter`): when set, the scan refuses any path that
    isn't under that root. Empty or unset → no containment, useful
    for single-user dev installs.
    """
    import os

    p = Path(root).expanduser()
    if not p.exists():
        return ScanResult(root=str(p), error=f"path does not exist: {p}")

    contain = os.environ.get("NICEMETA_FILE_SOURCE_ROOT", "").strip()
    if contain:
        contain_root = Path(contain).resolve()
        try:
            resolved = p.resolve()
            resolved.relative_to(contain_root)
        except (ValueError, OSError):
            return ScanResult(
                root=str(p),
                error=(
                    f"path is outside the configured file source root "
                    f"({contain_root})"
                ),
            )

    if p.is_file():
        # Treat a single file as a one-element folder so the UI flow stays
        # uniform — the user pastes a path and gets back an item list.
        return _classify_files([p], p.parent, max_files)

    if not p.is_dir():
        return ScanResult(root=str(p), error=f"path is neither file nor folder: {p}")

    paths: list[Path] = []
    if recursive:
        for entry in p.rglob("*"):
            if entry.is_file():
                paths.append(entry)
            if len(paths) > max_files * 2:
                # Safety: stop walking very large trees. We still classify
                # max_files worth so the user gets something useful.
                break
    else:
        paths = [entry for entry in p.iterdir() if entry.is_file()]

    return _classify_files(sorted(paths), p, max_files)


def _classify_files(paths: list[Path], root: Path, max_files: int) -> ScanResult:
    out = ScanResult(root=str(root))
    for fp in paths:
        if len(out.files) >= max_files:
            break
        ext = _extension_of(fp.name)
        fmt = _format_label(ext)
        if fmt == "unknown":
            out.skipped += 1
            continue
        try:
            size = fp.stat().st_size
        except OSError:
            size = 0
        rel = _safe_relative(fp, root)
        if fmt == "excel":
            sheets = list_excel_sheets(fp)
            # If we couldn't read sheets, surface the file as a single
            # entry — the user can still try to include it.
            if not sheets:
                out.files.append(ScannedFile(
                    uri=str(fp), name=fp.name, format="excel",
                    size_bytes=size, relative_path=rel,
                ))
                continue
            for sheet in sheets:
                out.files.append(ScannedFile(
                    uri=f"{fp}#{sheet}",
                    name=f"{fp.name} → {sheet}",
                    format="excel",
                    size_bytes=size,
                    relative_path=rel,
                    sheet=sheet,
                ))
        else:
            out.files.append(ScannedFile(
                uri=str(fp), name=fp.name, format=fmt,
                size_bytes=size, relative_path=rel,
            ))
    return out


def _safe_relative(fp: Path, root: Path) -> str:
    try:
        return str(fp.relative_to(root))
    except ValueError:
        return fp.name
